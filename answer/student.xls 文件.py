# encoding:utf-8

__author__ = 'APLK'
'''
纯文本文件 student.txt为学生信息, 里面的内容（包括花括号）如下所示：

{
	"1":["张三",150,120,100],
	"2":["李四",90,99,95],
	"3":["王五",60,66,68]
}

请将上述内容写到 student.xls 文件中，如下图所示：
'''
# workbook相关
# from openpyxl import Workbook
# from openpyxl.compat import range
# from openpyxl.utils import get_column_letter
#
# wb = Workbook()
#
# dest_filename = 'empty_book.xlsx'
#
# ws1 = wb.active
# ws1.title = "range names"
#
# for row in range(1, 10):
#     ws1.append(range(8))
#
# ws2 = wb.create_sheet(title="Pi")
#
# ws2['F5'] = 3.14
#
# ws3 = wb.create_sheet(title="Data")
# for row in range(5, 8):
#     for col in range(4, 7):
#        ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
# print(ws3['A8'].value)
# wb.save(filename=dest_filename)

from openpyxl import Workbook
from openpyxl.styles import Border, Side
# from  openpyxl.utils import get_column_letter
# from openpyxl.compat import range
import json
workbook = Workbook()
ws1 = workbook.active
ws1.title = "student"
with open('E:/python_demo/hello/answer/test/0014.txt','r', encoding = 'UTF-8') as f:
    read = f.read()
    data = json.loads(read, encoding='utf-8')
    row= len(data)
    print(data)
    i=0
    #设置边框样式，上下左右边框
    border = Border(left=Side(style='medium',color='FF000000'),right=Side(style='medium',color='FF000000'), bottom=Side(style='medium',color='FF000000'))
    for k,v in data.items():
        i+=1
        ws1.cell(row=i,column=1,value='{0}'.format(k)).border=border
        for n in range(2,len(v)+2):
            ws1.cell(row=i,column=n,value=v[n-2]).border=border
            # ws1.cell(row=i,column=n,value='{0}'.format(v[n-2]))
    ws1.cell(row=ws1.max_row+1,column=1,value='{0}'.format('TOTAL')).border=border
    ws1.cell(row=ws1.max_row,column=2,value='{0}'.format('NONE')).border=border
    ws1.cell(row=ws1.max_row,column=3,value='{0}'.format('=SUM(C1:C3)')).border=border
    ws1.cell(row=ws1.max_row,column=4,value='{0}'.format('=SUM(D1:D3)')).border=border
    ws1.cell(row=ws1.max_row,column=5,value='{0}'.format('=SUM(E1:E3)')).border=border
    workbook.save(filename='student.xlsx')


# import json
# with open('E:/python_demo/hello/answer/test/0014.txt','r', encoding = 'UTF-8') as f:
#     encode = f.read()
# load = json.loads(encode)
# worksheet = xlwt.Worksheet('student','E:/python_demo/hello/answer/test/student.xls', cell_overwrite_ok=False)
# for i in range(len(load)):
#     d = load[str(i+1)]
#     worksheet.write(i,0,i+1)
#     for j in range(len(d)):
#         worksheet.write(i,j+1,d[j])
# worksheet.save('E:/python_demo/hello/answer/test/student.xls')
# 进才认为招宝公司的股价在未来一个月会上涨，招宝公司目前股价是40元，
# 于是进才买入20张行权价为44元、一个月后到期的招宝公司认购期权（每张5000股）。
# 目前该期权价格为每股0.28元，进才共花费28,000元权利金。两周后，假设招宝公司的股价涨到45元，
# 此时期权的价格为每股1.7元。进才在平仓后获净利14.2=(1.7-0.28)*10万元，回报率为507%。